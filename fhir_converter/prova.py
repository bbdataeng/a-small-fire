"""
SCRIPT SPAZZATURA
might delete it later
"""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from typing import Dict
from loguru import logger as log

# wb = load_workbook(filename = r"C:\Users\Antonella\Desktop\FHIR\sample-colon-dataset.xlsx")


# header: Dict[int, str] = {}
# for ws in wb.worksheets:
#     if ws.sheet_state == "hidden":
#         log.warning("Ignoring hidden sheet: {}", ws.title)
#         continue

#     log.info("Reading sheet {}", ws.title)

#     for row in ws.iter_rows(min_row=2, max_row=2, max_col=999):
#         for cell in row:
            
#             if cell.value is None:
#                 break
#             header[cell.col_idx - 1] = cell.value
#     print(ws.max_row)
#     for row in ws.iter_rows(min_row=2, max_row=4, max_col=len(header)):
#         patient_data: Dict[str, str] = {}
#         # for idx, cell in enumerate(row):
#         for cell in row:
#             patient_data[header[cell.col_idx -1]] = cell.value

# print(header)
# print("PATIENT DATA --------------------")
# print(patient_data)




from fhir.resources.organization import Organization
from fhir.resources.address import Address
# json_str = '''{"resourceType": "Organization",
#                 "id": "f001",
#                 "active": True,
#                 "name": "Acme Corporation",     
#                 "address": [{"country": "Switzerland"}]
# }'''
# org = Organization.parse_raw(json_str)
# >>> isinstance(org.address[0], Address)
# True
# >>> org.address[0].country == "Switzerland"
# True
# >>> org.dict()['active'] is True
# True



# class Organization2(Organization):
#     _address: str

#     def __init__(self, _address: str, **kwargs: any) -> None:
#         self._address = _address
#         super().__init__(**kwargs)

#     class Config:
#         underscore_attrs_are_private = True

# data = {"resourceType": "Organization", "name": "Acme Corporation", "alias" : [],"id": "f001","active":True,
#         "_address": [{"country": "Switzerland"}]}
# #                 "id": "f001",
# #                 "active": True,
# #                 "name": "Acme Corporation",     
# #                 "address": [{"country": "Switzerland"}]
# # }
# org = Organization2(**data)

# print(org)



from fhir.resources.organization import Organization
from fhir.resources.address import Address

# add = Address(city="Rome", country="IT")
# print(add.city)

# data = {
#      "id": "f001",
#      "active": True,
#      "name": "Acme Corporation",
#      "contact" : "boh"
#  }
from pydantic import ValidationError

# Assuming you have already defined the Organization and ExtendedContactDetailType classes

# data = {
#     "id": "f001",
#     "active": True,
#     "name": "Acme Corporation",
#     "contact": [
#         {
#             "name": "John Doe",
#             "telecom": [
#                 {
#                     "system": "phone",
#                     "value": "123-456-7890",
#                     "use": "work"
#                 },
#                 {
#                     "system": "email",
#                     "value": "john.doe@example.com",
#                     "use": "work"
#                 }
#             ]
#         }
#     ]
# }

# try:
#     organization_instance = Organization(**data)
#     print(organization_instance.json(indent=2))
# except ValidationError as e:
#     print(f"Validation error: {e}")




# data = {
#     "id": "f001",
#     "active": True,
#     "name": "Acme Corporation",
#     "contact": [
#         {
#             "name": "John Doe",
#             "telecom": [
#                 {
#                     "system": "phone",
#                     "value": "123-456-7890",
#                     "use": "work"
#                 },
#                 {
#                     "system": "email",
#                     "value": "john.doe@example.com",
#                     "use": "work"
#                 }
#             ]
#         }
#     ]
# }
# try:
#     organization_instance = Organization(**data)
#     print(organization_instance.json(indent=2))
# except ValidationError as e:
#     print(f"Validation error: {e}")



## FUNZIONA:
# data = {
#     "id": "f001",
#     "active": True,
#     "name": "Acme Corporation",
#     "contact": [
#         {
#             "address": Address(city="Rome", country= "Italy"),
#             "telecom": [
#                 {
#                     "system": "phone",
#                     "value": "123-456-7890",
#                     "use": "work"
#                 },
#                 {
#                     "system": "email",
#                     "value": "john.doe@example.com",
#                     "use": "work"
#                 }
#             ]
#         }
#     ]
# }
# try:
#     organization_instance = Organization(**data)
#     print(organization_instance.json(indent=2))
# except ValidationError as e:
#     print(f"Validation error: {e}")



org = Organization(
            id="IDID",
            active=True,
            # Name used for the organization
            name="Istituto Nazionale Tumori Regina Elena",
            # A list of alternate names that the organization is known as,
            # or was known as in the past
            alias=["IRE"],
            # A contact detail for the organization
            contact= [
        {
            "address": Address(use="work",
                        type="physical",
                        country="IT",
                        city="Rome",
                        postalCode="00144",
                        line=["Via Chianesi 53"]),
            "telecom": [
                {
                    "system": "phone",
                    "value": "+390652666931",
                    "use": "work",
                    "rank": 1
                },
                {
                    "system": "email",
                    "value": "laura.conti@ifo.gov.it",
                    "use": "work",
                    "rank": 1
                }
            ]
        }]
        )
print(org)

# org = Organization(**data)
# org.resource_type == "Organization"
# print(org)

# isinstance(org.address[0], Address)

# org.address[0].country == "Switzerland"

# org.dict()['active'] is True





from fhir_biobank.patient import PatientResource
from fhir_biobank.bundle import Bundle, Entry
from datetime import date

internal_id = "0"
patient_identifier = "4816522"
patient_gender = "female"
patient_birthdate = date(2000, 12, 11)
patient = PatientResource(internal_id, patient_identifier, patient_gender, patient_birthdate)

print("PATIENT:\n", patient.patientJSON())

fullURL_patient_resource = "https://example.com/patient/0"
shortURL_patient_resource = "patient/0"
entry = Entry(patient, fullURL_patient_resource, shortURL_patient_resource)


print("ENTRY:\n", entry.entryJSON())

entries = [entry]
bundle_id = "424242"
bundle = Bundle(bundle_id, entries)

print("BUNDLE:\n", bundle.bundleJSON())


from fhir_biobank.diagnosis import Diagnosis

diagnos = Diagnosis("C18.1")
print(diagnos)