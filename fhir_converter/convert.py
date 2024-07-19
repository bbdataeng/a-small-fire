# py convert.py --filename "C:\\Users\\Antonella\\Desktop\\FHIR\\sample-colon-dataset.xlsx" --outdir "C:\\Users\\Antonella\\Desktop\\FHIR"
import sys
from pathlib import Path
from typing import Dict
import os
import simplejson as json  # base json is unable to properly serialize Decimals
import typer
from fhir_model import FHIRSerializer, bbmri_post_serialization
from fhir_resources import FHIRResources
from input_models import Patient as PatientInputModel
from loguru import logger as log
from normalization import normalize_input, normalize_output
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

# from tqdm import tqdm


"""
conversion.py -----------------------------------------------------------------
overall wrapper, takes as input one excel (classic dataset table) written in 
the BBMRI appendix format and creates the FHIR-structured JSON to be uploaded 
on the Biobank Locator

-------------------------------------------------------------------------------

Minimal denominators' findable in the Federated Platform tools:
    # ########## Donor/Clinical Information
    # Patient pseudonym
    PATIENT_ID: Optional[str]
    # Age at diagnosis (rounded to years)
    AGE_AT_PRIMARY_DIAGNOSIS: int
    # Biological sex
    SEX: SEX_ENUM
    # Date of diagnosis
    DATE_DIAGNOSIS: date
    # Diagnosis (ICD-10)
    DIAGNOSIS : DIAGNOSIS_ENUM
    # Donor Age / Date of birth
    DONOR_AGE : date
    # AGE: int

    # ########## Sample
    # Sample ID
    SAMPLE_ID: str
    # Material type
    SAMPLE_MATERIAL_TYPE: SAMPLE_MATERIAL_TYPE_ENUM
    # Year of sample collection
    YEAR_OF_SAMPLE_COLLECTION: int
    # Storage temperature
    STORAGE_TEMPERATURE: Optional[str]

"""


app = typer.Typer()


@app.command()
def convert(
    filename: Path = typer.Option(..., help="Path of input file"),
    outdir: Path = typer.Option(..., help="Path of output folder"),
    miabis: bool = typer.Option(default= False, help= "Flag for MIABIS normalization"),
) -> None:

    if not filename.exists():
        sys.exit(f"ERROR: File '{filename}' does not exist")

    if not outdir.exists():
        sys.exit(f"ERROR: Outdir '{outdir}' does not exist")

    try:
        wb = load_workbook(filename=filename)
    except InvalidFileException as e:
        sys.exit(str(e))

    header: Dict[int, str] = {}

    # create a bundle with collection resource and biobank resource
    bundle = FHIRResources.get_bundle()

    collection = FHIRResources.get_organization("Collection")
    biobank = FHIRResources.get_organization("Biobank")
    bundle.entry.append(collection)
    bundle.entry.append(biobank)
    bundle_data = bundle.dict()
    bundle_data = bbmri_post_serialization(bundle_data)

    ## create a json file for the organization
    with open(f"{outdir}/organization.json", "w") as f:
        json.dump(bundle_data, f, default=str, indent=4)

    for ws in wb.worksheets:
        if ws.sheet_state == "hidden":
            log.warning("Ignoring hidden sheet: {}", ws.title)
            continue

        log.info("Reading sheet {}", ws.title)

        ## create header
        ## for each cell save the value (header name), until None (columns end)
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=999):
            for cell in row:
                if cell.value is None:
                    break
                header[cell.col_idx - 1] = cell.value

        ## CREATE PATIENT DATA
        bundle = FHIRResources.get_bundle()
        patients_ids = []
        counters: Dict[str, int] = {}
        ## for each cell, take the value and put it in patient_data dict with key = header
        # for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(header)):
        for row_number, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(header)), start=2
        ):
            if ws.row_dimensions[row_number].hidden:  # if the row is hidden, skip it
                continue

            patient_data: Dict[str, str] = {}

            for cell in row:
                patient_data[header[cell.col_idx - 1]] = cell.value

            if all(
                value is None for value in patient_data.values()
            ):  # if all the fields are None --> stop (no more rows)
                break

            # normalize_input: make input fields MIABIS compliant
            if not miabis: # if data is not MIABIS compliant 
                # patient_data = normalize_input(patient_data)  # aggiungere Flag!
                patient_data = normalize_input(patient_data, "field_mapping_config.yml")

            try:
                ## Create Patient Resource with patient_data
                ## validation of fields with pydantic --> check MIABIS compliance
                patient = PatientInputModel(**patient_data)
            except ValidationError as e:
                print(e)

                for field in str(e).split("\n")[1::2]:
                    log.error(
                        "{}: {} = [{}]",
                        patient_data.get("SAMPLE_ID", ""),
                        field,
                        patient_data.get(field, "N/A"),
                    )
            else:
                patient = normalize_output(patient)  # mapping to BBMRI implementation guide
                # print("PATIENT\n", patient)
                patient_serializer = FHIRSerializer(patient, counters) # mapping to FHIR resources

                pat_id = patient_serializer.PATIENT_ID
                copy = pat_id in patients_ids

                patients_ids.append(pat_id)

                sample_id, bundle = patient_serializer.serialize_patient(bundle, copy)
                bundle_data = bundle.dict()
                bundle_data = bbmri_post_serialization(bundle_data)

                with open(f"{outdir}/bundle-{bundle.id}.json", "w") as f:
                    json.dump(bundle_data, f, default=str, indent=4)

        # Parse the first sheet only
        break


if __name__ == "__main__":
    app()
